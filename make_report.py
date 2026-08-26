# -*- coding: utf-8 -*-
# 生成《できる日本語 初級》单词卡程序 vs 官方词汇表 校验报告 Excel
import json, re, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 复用 compare.py 的 PDF 词汇表
ns = {}
exec(open(r'D:\できる日本語 单词卡\compare.py', encoding='utf-8').read().split('app = json.load')[0], ns)
pdf_words = ns['pdf_words']
def norm(s):
    if s is None: return ''
    s = s.replace('~','～').replace(' ','').replace('　','').rstrip('。．.')
    s = re.sub(r'[（(][^）)]*[）)]', '', s)
    return s

app = json.load(open(r'D:\できる日本語 单词卡\app_words.json', encoding='utf-8'))

# 每课：缺失词（原文，去掉句子类标注时保留）、程序独有
report = {}
for k in sorted(pdf_words, key=int):
    pdf_items = pdf_words[k]
    pdfset = set(norm(w) for w in pdf_items)
    appset = set(norm(w['kanji']) for w in app[str(k)])
    missing_raw = []
    for w in pdf_items:
        if norm(w) not in appset:
            missing_raw.append(w)
    extra_raw = []
    for w in app[str(k)]:
        if norm(w['kanji']) not in pdfset:
            extra_raw.append(w['kanji'])
    report[k] = {'pdf': len(pdfset), 'app': len(appset), 'missing': missing_raw, 'extra': extra_raw}

wb = Workbook()
thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
head_fill = PatternFill('solid', fgColor='4F46E5')
head_font = Font(color='FFFFFF', bold=True, size=11)
warn_fill = PatternFill('solid', fgColor='FFF3CD')
ok_fill = PatternFill('solid', fgColor='D4EDDA')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ---------- Sheet1 汇总 ----------
ws = wb.active
ws.title = '汇总'
ws.append(['课', '官方词汇表词条数', '程序词条数', '缺失词条数', '缺失率', '覆盖情况', '备注'])
for c in ws[1]:
    c.fill = head_fill; c.font = head_font; c.alignment = center; c.border = border
lessons = ['第1课 はじめまして','第2课 買い物・食事','第3课 スケジュール','第4课 私の国・町',
           '第5课 休みの日','第6课 一緒に!','第7课 友達の家で','第8课 大切な人','第9课 好きなこと',
           '第10课 バスツアー','第11课 私の生活','第12课 病気・けが','第13课 私のおすすめ',
           '第14课 国・町の習慣','第15课 イベント情報・ニュースから']
total_miss = 0
for idx, k in enumerate(sorted(report, key=int), start=2):
    r = report[k]
    miss = len(r['missing'])
    total_miss += miss
    rate = miss / r['pdf'] * 100
    note = ''
    if k == 5:
        note = '程序第5课为另选词表，与官方词汇表差异最大，含表外词(連休/遅刻/試験等)'
    elif k == 15 and miss == 0:
        note = '完全覆盖官方词汇表'
    if r['extra']:
        note += ('；' if note else '') + '含程序独有词{}个'.format(len(r['extra']))
    status = '完整' if miss == 0 else '有缺失'
    ws.append([lessons[idx-2], r['pdf'], r['app'], miss, '{:.0f}%'.format(rate), status, note])
    for col in range(1, 8):
        cell = ws.cell(row=idx, column=col)
        cell.border = border
        cell.alignment = left if col in (1,7) else center
        if miss == 0:
            cell.fill = ok_fill
        elif col == 5 and rate > 50:
            cell.fill = warn_fill
# 合计行
tot_row = len(report) + 2
ws.append(['合计', sum(r['pdf'] for r in report.values()), sum(r['app'] for r in report.values()), total_miss,
           '{:.0f}%'.format(total_miss/sum(r['pdf'] for r in report.values())*100), '有缺失', ''])
for col in range(1, 8):
    c = ws.cell(row=tot_row, column=col); c.border = border; c.font = Font(bold=True)
    c.alignment = left if col in (1,7) else center
widths = [22, 18, 12, 12, 10, 12, 60]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ---------- Sheet2 缺失词清单 ----------
ws2 = wb.create_sheet('缺失词清单')
ws2.append(['课', '缺失的官方词汇', '类型说明'])
for c in ws2[1]:
    c.fill = head_fill; c.font = head_font; c.alignment = center; c.border = border
row = 2
for k in sorted(report, key=int):
    r = report[k]
    if not r['missing']:
        continue
    ws2.append([lessons[k-1], '', ''])
    ws2.cell(row=row, column=1).font = Font(bold=True)
    ws2.cell(row=row, column=1).fill = PatternFill('solid', fgColor='E8E8FF')
    ws2.cell(row=row, column=1).border = border
    ws2.cell(row=row, column=2).border = border
    ws2.cell(row=row, column=3).border = border
    row += 1
    for w in r['missing']:
        typ = '句子/寒暄' if any(ch in w for ch in 'です。ます。') else ('量词' if w.startswith('～') else '词条')
        ws2.append(['', w, typ])
        ws2.cell(row=row, column=2).alignment = left
        ws2.cell(row=row, column=3).alignment = center
        for col in range(1, 4):
            ws2.cell(row=row, column=col).border = border
        row += 1
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 42
ws2.column_dimensions['C'].width = 14

# ---------- Sheet3 程序独有词 ----------
ws3 = wb.create_sheet('程序独有词')
ws3.append(['课', '程序独有词汇', '说明'])
for c in ws3[1]:
    c.fill = head_fill; c.font = head_font; c.alignment = center; c.border = border
row = 2
for k in sorted(report, key=int):
    r = report[k]
    if not r['extra']:
        continue
    for w in r['extra']:
        note = ''
        if w in ('一昨日','明後日'):
            note = '官方表用假名(おととい/あさって)，程序用汉字，本质同词'
        elif w in ('会う','寝る','休む','遊ぶ','買い物','食事'):
            note = '官方表用ます形(会います/寝ます/休みます/遊びます/買い物します/食事します)，程序用辞书形'
        elif w in ('連休','遅刻','試験'):
            note = '不在官方词汇表，为教材正文/补充词'
        elif w == 'お金':
            note = '官方表作(お)金，本质同词'
        elif w in ('入学','卒業'):
            note = '官方表作入学します/卒業します，程序用名词形'
        elif w in ('びっくり','化粧','留学','経験'):
            note = '官方表作〜します，程序用名词形'
        else:
            note = '与官方词条为同词异写'
        ws3.append([lessons[k-1], w, note])
        for col in range(1, 4):
            ws3.cell(row=row, column=col).border = border
            ws3.cell(row=row, column=col).alignment = left if col == 3 else center
        row += 1
ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 60

# ---------- Sheet4 程序数据检查 ----------
ws4 = wb.create_sheet('程序数据检查')
ws4.append(['课', '词汇', '字段', '当前值', '应为', '说明'])
for c in ws4[1]:
    c.fill = head_fill; c.font = head_font; c.alignment = center; c.border = border
issues = [
    (4, '東', 'kana', 'हिगाशि', 'ひがし', '假名位置误填尼泊尔语/天城文'),
    (12, '歯', 'kana', 'ह', 'は', '假名位置误填天城文'),
    (15, 'かかります', 'kana', 'काकरु', 'かかります', '假名位置误填天城文'),
]
row = 2
for it in issues:
    ws4.append(list(it))
    for col in range(1, 7):
        ws4.cell(row=row, column=col).border = border
        ws4.cell(row=row, column=col).alignment = left if col in (3,4,5,6) else center
    row += 1
ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 8
ws4.column_dimensions['D'].width = 14
ws4.column_dimensions['E'].width = 14
ws4.column_dimensions['F'].width = 40

out = r'D:\できる日本語 单词卡\单词卡词汇完整性校验报告.xlsx'
wb.save(out)
print('saved:', out)
print('汇总缺失总数:', total_miss)
for k in sorted(report, key=int):
    r = report[k]
    print('第{}课: PDF {} | 程序 {} | 缺失 {}'.format(k, r['pdf'], r['app'], len(r['missing'])))
