# -*- coding: utf-8 -*-
# 生成中级三语词汇表 Excel + 飞书sheets结构JSON
import json, sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE = r'/Users/kanryo/Documents/工作空间/できる日本語_单词卡_项目'
rows = json.load(open(os.path.join(BASE, 'med_trilingual.json'), encoding='utf-8'))

TITLES = {
 1:'新しい出会い',2:'楽しい食事・上手な買い物',3:'時間を生かす',4:'地域を知って生活する',
 5:'緊急事態！',6:'地図を広げる',7:'世代を超えた交流',8:'気持ちを伝える',9:'言葉を楽しむ',
 10:'日本を旅する',11:'ライフスタイル',12:'心と体の健康',13:'トレンドに乗ってつながる',
 14:'カルチャーショック',15:'地域社会に生きる',16:'学校生活',17:'働くということ',
 18:'地球に生きる',19:'科学の力',20:'豊かさと幸せ'
}

def ne_val(r):
    v = (r.get('ne') or '').strip()
    if not v:
        return '（公式リストになし）'
    return v

# ---------- Excel ----------
wb = Workbook()
ws = wb.active
ws.title = '全20課一覧'
headers = ['課','番号','単語','よみ','英語','中国語','ネパール語']
hfill = PatternFill('solid', fgColor='4F46E5')
hfont = Font(color='FFFFFF', bold=True, size=12)
title_fill = PatternFill('solid', fgColor='EEF2FF')
thin = Border(*[Side(style='thin', color='C7D2FE')]*4)
for c,h in enumerate(headers,1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = hfill; cell.font = hfont
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin
ws.freeze_panes = 'A2'

r = 2
for lesson in range(1,21):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    tc = ws.cell(row=r, column=1, value='第%d課　%s' % (lesson, TITLES[lesson]))
    tc.fill = title_fill; tc.font = Font(bold=True, size=13, color='312E81')
    tc.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(1,8):
        ws.cell(row=r, column=c).border = thin
    r += 1
    for item in rows:
        if item['lesson'] != lesson: continue
        vals = ['第%d課'%lesson, item['no'], item['kanji'], item['kana'], item['en'], item['cn'], ne_val(item)]
        for c,v in enumerate(vals,1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin
            cell.font = Font(size=11)
        r += 1

widths = [9,6,30,22,44,34,40]
for i,w in enumerate(widths,1):
    ws.column_dimensions[chr(64+i)].width = w
ws.row_dimensions[1].height = 24
wb.save(os.path.join(BASE, '中级词汇表.xlsx'))
print('Excel 已保存: 中级词汇表.xlsx')

# ---------- 飞书sheets结构 ----------
payload = {'sheets': [{'name':'全20課一覧','columns':headers,'data':[]}]}
for lesson in range(1,21):
    payload['sheets'][0]['data'].append(['第%d課　%s'%(lesson,TITLES[lesson]),'','','','','',''])
    for item in rows:
        if item['lesson'] != lesson: continue
        payload['sheets'][0]['data'].append(['第%d課'%lesson, item['no'], item['kanji'], item['kana'], item['en'], item['cn'], ne_val(item)])
with open(os.path.join(BASE,'med_trilingual_sheet.json'),'w',encoding='utf-8') as f:
    json.dump(payload, f, ensure_ascii=False, indent=1)
print('sheets JSON 已保存: med_trilingual_sheet.json, 共 %d 行' % len(payload['sheets'][0]['data']))

# ---------- 单词卡数据（app_words格式）----------
app = {}
for lesson in range(1,21):
    arr = []
    for item in rows:
        if item['lesson'] != lesson: continue
        arr.append({'kanji':item['kanji'],'kana':item['kana'],'eng':item['en'],'cn':item['cn'],'ne':ne_val(item)})
    app[str(lesson)] = arr
with open(os.path.join(BASE,'med_app_words.json'),'w',encoding='utf-8') as f:
    json.dump(app, f, ensure_ascii=False)
print('单词卡数据已保存: med_app_words.json')
print('各课词数:', {k:len(v) for k,v in app.items()})
